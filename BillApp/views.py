from django.shortcuts import render, redirect
from django.contrib.auth.models import User, auth
from django.contrib import messages
from random import randint
from django.core.mail import send_mail, EmailMessage
from io import BytesIO
from django.db import transaction
from django.conf import settings
from django.db import connection
from datetime import datetime, date, timedelta
# import requests
from decimal import Decimal
from num2words import num2words
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import *
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Protection, Alignment
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from collections import defaultdict
from django.db.models import Sum

# Create your views here.



def is_admin(user):
    return user.groups.filter(name="ADMIN").exists()

# @login_required(login_url='/')
# @user_passes_test(is_admin, login_url='/')
# def goAdminPanel(request):
#     try:
#         return render(request, 'admin/admin_index.html')
#     except PermissionDenied:
#         messages.error(request, 'You are not allowed to access this page!')
#         return redirect('/')
#     except Exception as e:
#         print(e)
#         messages.error(request, 'An error occurred while processing your request.')
#         return redirect('/')

# def goAdminPanel(request):
#     try:
#         return render(request, 'admin/admin_index.html')
#     except Exception as e:
#         print(e)
#         messages.error(request, 'An error occurred while processing your request.')
#         return redirect('/')

#Admin Panel
def goRegisteredClients(request):
    if request.user.is_staff:
        all_companies = Company.objects.all()
        clients = []
        for i in all_companies:
            try:
                trial = ClientTrials.objects.filter(company = i).first()
            except:
                trial = None
            dict = {'company':i,'trial':trial}
            clients.append(dict)
        context = {
            'clients' : clients,
        }
        return render(request, 'admin/reg_clients.html',context)
    else:
        return('/')


def goDemoClients(request):
    if request.user.is_staff:
        context = {
            # 'clients' : ClientTrials.objects.filter(trial_status = True),
            'clients': ClientTrials.objects.filter(Q(trial_status=True) | (Q(trial_status=False) & Q(subscribe_status='yes'))),
            'terms' : PaymentTerms.objects.all()
        }
        return render(request, 'admin/demo_clients.html',context)
    else:
        return redirect('/')

def goPurchasedClients(request):
    if request.user.is_staff:
        context = {
            'clients' : ClientTrials.objects.exclude(purchase_status = 'null')
        }
        return render(request, 'admin/purchased_clients.html',context)
    else:
        return redirect('/')

def cancelSubscription(request,id):
    if request.user.is_staff:
        status = ClientTrials.objects.get(id = id)
        status.purchase_status = 'cancelled'
        status.save()
        messages.success(request, 'Subscription Cancelled.!')
        return redirect(goPurchasedClients)
    return redirect('/')

def goPaymentTerms(request):
    if request.user.is_staff:
        terms = PaymentTerms.objects.all()
        return render(request, 'admin/payment_terms.html',{'terms':terms})
    return redirect('/')

def addNewPaymentTerm(request):
    if request.user.is_staff:
        return render(request, 'admin/add_payment_term.html')
    else:
        return redirect('/')
    
def createPaymentTerm(request):
    if request.user.is_staff:
        if request.method == 'POST':
            dur = request.POST['duration']
            term = request.POST['term']
            dys = int(dur) if term == 'Days' else int(dur) * 30

            PaymentTerms.objects.create(duration = dur, term = term, days = dys)
            messages.success(request, 'Success.!')
            
            if 'next_term' in request.POST:
                return redirect(addNewPaymentTerm)
            else:
                return redirect(goPaymentTerms)
    else:
        return redirect('/')

def deletePaymentTerm(request, id):
    if request.user.is_staff:
        term = PaymentTerms.objects.get(id = id)
        term.delete()
        return redirect(goPaymentTerms)
    return redirect('/')


def clientPurchase(request, id):
    if request.user.is_staff:
        client = ClientTrials.objects.get(id = id)

        if request.method == 'POST':
            start = request.POST['purchaseDate']
            end = request.POST['endDate']
            term = PaymentTerms.objects.get(id = request.POST['paymentTerm'])

            client.purchase_start_date = start
            client.purchase_end_date = end
            client.payment_term = str(term.duration)+" "+term.term
            client.purchase_status = 'valid'
            client.trial_status = False
            client.subscribe_status = 'purchased'
            client.save()

            messages.success(request,'Success.!')
            return redirect(goDemoClients)
        return redirect(goDemoClients)
    return redirect('/')


def getPaymentTerms(request):
    if request.user.is_staff:
        try:
            terms = PaymentTerms.objects.all()
            list = []

            for item in terms:
                paymentTerms = {
                    "id": item.id,
                    "days": item.days,
                    "term": item.term,
                    "duration" : item.duration,
                }
                list.append(paymentTerms)

            print(list)
            return JsonResponse({"terms": list}, safe=False)
        except Exception as e:
            print(e)
            return JsonResponse({"message": "failed"})
    else:
        return JsonResponse({"message": "failed"})

def removeUser(request,id):
    if request.user.is_staff:
        try:
            usr = User.objects.get(id = id)
            usr.delete()
            return redirect(goRegisteredClients)
        except Exception as e:
            print(e)
            return redirect(goRegisteredClients)
    else:
        return redirect('/')


def index(request):
    return render(request, "index.html")

def goBlog(request):
    return render(request, "blog.html")

def login(request):
    return render(request, "login.html")


# Conversion long num to String mode
def human_format(num):
    magnitude = 0
    while abs(num) >= 1000:
        magnitude += 1
        num /= 1000.0
    return "%.2f%s" % (num, ["", "K", "M", "G", "T", "P"][magnitude])


@login_required(login_url="login")
def goDashboard(request):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        try:
            sales = Sales.objects.filter(cid = cmp)
            purchases = Purchases.objects.filter(cid = cmp)
            items = Items.objects.filter(cid = cmp)
            todSale = 0
            totSale = 0
            todPurchase=0
            totPurchase = 0
            for i in sales:
                totSale += i.total_amount
                if i.date == date.today():
                    todSale += i.total_amount

            for i in purchases:
                totPurchase += i.total_amount
                if i.date == date.today():
                    todPurchase += i.total_amount


            # Chart data 

            data1 = []
            data2 = []
            data3 = []
            data4 = []
            data5 = []
            label = []
            
            for yr in range((date.today().year)-4, (date.today().year)+1):
                label.append(yr)
                salesAmount = 0
                purchaseAmount = 0
                for i in sales:
                    if i.date.year == yr:
                        salesAmount+= i.total_amount

                for i in purchases:
                    if i.date.year == yr:
                        purchaseAmount += i.total_amount

                data1.append(float(salesAmount))
                data2.append(float(purchaseAmount))

                stockIn = 0
                stockOut = 0
                stockBalance = 0
                for i in Item_transactions.objects.filter(cid = cmp).filter(type = 'Purchase'):
                    if i.date.year == yr:
                        stockIn += i.quantity

                for i in Item_transactions.objects.filter(cid = cmp).filter(type = 'Sale'):
                    if i.date.year == yr:
                        stockOut += i.quantity

                for i in Item_transactions.objects.filter(cid = cmp):
                    if i.date.year == yr and (i.type == "Opening Stock" or i.type == 'Add Stock' or i.type == 'Purchase'):
                            stockBalance += i.quantity
                    
                    if i.date.year == yr and (i.type == "Reduce Stock" or i.type == 'Sale'):
                            stockBalance -= i.quantity
                            
                data3.append(stockIn)
                data4.append(stockOut)
                data5.append(stockBalance)

            context = {
                "cmp": cmp,
                "todSale": f"{todSale:.2f}",
                "totSale": f"{totSale:.2f}",
                "todPurchase": f"{todPurchase:.2f}",
                "totPurchase": f"{totPurchase:.2f}",
                'salesData':data1,
                'purchaseData':data2,
                'stockIn':data3,
                'stockOut':data4,
                'stockBalance':data5,
                'label':label,

            }
            return render(request, "dashboard.html", context)
        except Exception as e:
            print(e)
            return redirect('/')
    


def redirectPage(request):
    try:
        search = str(request.GET['url']).lower()
        if search in ('sale','sales','add sales','sales add','create sales','sales create', 'new sale', 'sale new'):
            return redirect(addNewSale)
        elif search in ('purchase','add purchase','purchase add','create purchase','purchase create', 'new purchase', 'purchase new'):
            return redirect(addNewPurchase)
        elif search in ('stock','stock reports','reports'):
            return redirect(goStockReports)
        elif search in ('item','add item','new item', 'item add'):
            return redirect(addNewItem)
        else:
            messages.error(request, "Not Found.!")
            return redirect(goDashboard)
    except Exception as e:
        print(e)
        return redirect(goDashboard)


def registerUser(request):
    try:
        if request.method == "POST":
            usrnm = request.POST["username"]
            eml = request.POST["email"]
            phn = request.POST["phone"]
            adrs = request.POST["address"]
            gstn = request.POST["gstnum"]
            cmpny = request.POST["company"]
            state = request.POST['state']
            cntry = request.POST['country']
            pswrd = request.POST["password"]
            cpswrd = request.POST["confirmPassword"]

            if User.objects.filter(username=usrnm).exists():
                messages.info(
                    request, f"`{usrnm}` already exists!! Please Login or try another.."
                )
                return redirect(login)
            elif User.objects.filter(email=eml).exists():
                messages.info(request, f"`{eml}` already exists!! Please try another..")
                return redirect(login)
            elif Company.objects.filter(phone_number = phn).exists():
                messages.info(request, f"Phone number already exists!! Please try another..")
                return redirect(login)
            elif Company.objects.filter(company_name__iexact=cmpny.lower()).exists():
                messages.info(request, f"Company Name `{cmpny}` already exists!! Please try another..")
                return redirect(login)
            else:
                if pswrd == cpswrd:
                    userInfo = User.objects.create_user(
                        username=usrnm,
                        email=eml,
                        password=pswrd,
                    )
                    userInfo.save()
                    print("auth user saved...")
                    cData = User.objects.get(id=userInfo.id)
                    cmpnyData = Company(
                        user=cData,
                        company_name=cmpny,
                        phone_number=phn,
                        address=adrs,
                        gst_number=gstn,
                        state = state,
                        country = cntry,
                    )
                    cmpnyData.save()
                    
                    #storing trail data
                    start = date.today()
                    end = start + timedelta(days=30)
                    trail = ClientTrials(
                        user = cData,
                        company = cmpnyData,
                        start_date = start,
                        end_date = end,
                        trial_status = True,
                        purchase_start_date = None,
                        purchase_end_date = None,
                        purchase_status = "null",
                        payment_term = None,
                        subscribe_status = 'null',
                    )
                    trail.save()

                    messages.info(request, 'Registration Successful..')
                    return redirect(login)
                else:
                    messages.warning(request, "Passwords doesn't match..Please try again.")
                    # return HttpResponse('please! verify your passwords')
                    return redirect(login)
        else:
            return redirect(login)
    except Exception as e:
        print(e)
        return redirect(login)

def registerTrialUser(request):
    try:
        if request.method == "POST":
            usrnm = request.POST["username"]
            eml = request.POST["email"]
            phn = request.POST["phone"]
            adrs = request.POST["address"]
            gstn = request.POST["gstnum"]
            cmpny = request.POST["company"]
            state = request.POST['state']
            cntry = request.POST['country']
            pswrd = request.POST["password"]
            cpswrd = request.POST["confirmPassword"]

            if User.objects.filter(username=usrnm).exists():
                res = f'<script>alert("User name `{usrnm}` already exists, Please Login or try another.!");window.history.back();</script>'
                return HttpResponse(res)
            elif User.objects.filter(email=eml).exists():
                res = f'<script>alert("Email `{eml}` already exists!! Please try another..");window.history.back();</script>'
                return HttpResponse(res)
            elif Company.objects.filter(phone_number = phn).exists():
                res = f'<script>alert("Phone number already exists!! Please try another..");window.history.back();</script>'
                return HttpResponse(res)
            elif Company.objects.filter(company_name__iexact=cmpny.lower()).exists():
                res = f'<script>alert("Company Name `{cmpny}` already exists!! Please try another..");window.history.back();</script>'
                return HttpResponse(res)
            else:
                if pswrd == cpswrd:
                    userInfo = User.objects.create_user(
                        username=usrnm,
                        email=eml,
                        password=pswrd,
                    )
                    userInfo.save()
                    print("auth user saved...")
                    cData = User.objects.get(id=userInfo.id)
                    cmpnyData = Company(
                        user=cData,
                        company_name=cmpny,
                        phone_number=phn,
                        address=adrs,
                        gst_number=gstn,
                        state = state,
                        country = cntry,
                    )
                    cmpnyData.save()

                    #storing trial data
                    start = date.today()
                    end = start + timedelta(days=30)
                    trial = ClientTrials(
                        user = cData,
                        company = cmpnyData,
                        start_date = start,
                        end_date = end,
                        trial_status = True,
                        purchase_start_date = None,
                        purchase_end_date = None,
                        purchase_status = "null",
                        payment_term = None,
                        subscribe_status = 'null',
                    )
                    trial.save()

                    messages.success(request, 'Registration Successful..')
                    return redirect(login)
                else:
                    # messages.warning(request, "Passwords doesn't match..Please try again.")
                    # return redirect(login)
                    res = f'<script>alert("Passwords does not match..Please try again..");window.history.back();</script>'
                    return HttpResponse(res)
        else:
            return redirect('/')
    except Exception as e:
        print(e)
        return redirect('/')

def userLogin(request):
    if request.method == "POST":
        uName = request.POST["username"]
        password = request.POST["password"]

        user = auth.authenticate(username=uName, password=password)
        if user is not None:
            if user.is_staff:
                auth.login(request, user)
                return redirect(goRegisteredClients)
            else:
                status = ClientTrials.objects.get(user = user.id)
                if status.purchase_status == 'valid':
                    auth.login(request, user)
                    return redirect(goDashboard)
                elif status.purchase_status == 'expired':
                    messages.warning(request, "Your Subscription has been expired.! Contact Admin.")
                    return redirect(login)
                elif status.purchase_status == 'cancelled':
                    messages.warning(request, "Your Subscription has been Cancelled.! Contact Admin.")
                    return redirect(login)
                else:
                    if status.trial_status:
                        auth.login(request, user)
                        return redirect(goDashboard)
                    else:
                        messages.warning(request, "Your Trial period has been expired.! Contact Admin.")
                        return redirect(login)
        else:
            messages.info(request, "Incorrect Username or Password..Please try again")
            return redirect(login)
    else:
        return redirect(login)


def showProfile(request):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        try:
            context = {
                'cmp':cmp,
            }
            return render(request, 'profile.html',context)
        except Exception as e:
            print(e)
            return redirect("/")
    return redirect("/")


@login_required(login_url="login")
def updateUserProfile(request):
    if request.user:
        user = User.objects.get(id = request.user.id)
        cmp = Company.objects.get(user = user.id)
        try:
            if request.method == 'POST':
                cmp.company_name = request.POST['company_name']
                cmp.gst_number = request.POST['gst_number']
                cmp.phone_number = request.POST['phone_number']
                cmp.address = request.POST['address']
                cmp.state = request.POST['state']
                cmp.country = request.POST['country']
                cmp.save()
            
                if user.username != request.POST['username'] and User.objects.filter(username = request.POST['username']).exists():
                    messages.error(request, 'Username already exists, Try another.!')
                    return redirect(showProfile)
                if user.username != request.POST['username'] and User.objects.filter(email = request.POST['email']).exists():
                    messages.error(request, 'Email already exists, Try another.!')
                    return redirect(showProfile)
                    
                user.username = request.POST['username']
                user.email = request.POST['email']
                user.save()
                
                messages.success(request, 'Profile updated successfully.!')
                return redirect(showProfile)
        except Exception as e:
            print(e)
            return redirect(showProfile)
    return redirect('/')


def forgotPassword(request):
    try:
        email = request.POST['email']
        user = User.objects.filter(email = email).first()
        if User.objects.filter(email = email).exists():
            password = str(randint(100000, 999999))
            # print(password)
            user.set_password(password)
            user.save()

            # SEND MAIL CODE
            subject = "Forgot Password"
            message = f"Dear user,\nYour Password has been reset as you requested. You can login with the password given below\n\nPassword:{password}"
            recipient = user.email
            send_mail(subject, message, settings.EMAIL_HOST_USER, [recipient])


            return JsonResponse({'message':'success'})
        else:
            return JsonResponse({'message':'not_found'})
    except Exception as e:
        print(e)
        return redirect(login)
    

@login_required(login_url="login")
def updateLogo(request,id):
    if request.user:
        cmp = Company.objects.get(user = id)
        try:
            if request.method == 'POST':
                cmp.logo = request.FILES.get('logo')
                cmp.save()
                return redirect(showProfile)
        except Exception as e:
            print(e)
            return redirect(showProfile)
    return redirect('/')


@login_required(login_url="login")
def removeLogo(request):
    if request.user:
        cmp = Company.objects.get(user = request.user.id)
        try:
            cmp.logo = None
            cmp.save()
            return redirect(showProfile)
        except Exception as e:
            print(e)
            return redirect(showProfile)
    return redirect('/')



@login_required(login_url="login")
def userLogout(request):
    request.session["uid"] = ""
    auth.logout(request)
    return redirect("/")


def validateEmail(request):
    email = request.GET["email"]

    if User.objects.filter(email=email).exists():
        return JsonResponse({"is_taken": True})
    else:
        return JsonResponse({"is_taken": False})


def validateUsername(request):
    uName = request.GET["username"]

    if User.objects.filter(username=uName).exists():
        return JsonResponse({"is_taken": True})
    else:
        return JsonResponse({"is_taken": False})

def validatePhone(request):
    number = request.GET["phone"]

    if Company.objects.filter(phone_number=number).exists():
        return JsonResponse({"is_taken": True})
    else:
        return JsonResponse({"is_taken": False})


def validateCompany(request):
    cmp = request.GET["company"]

    if Company.objects.filter(company_name__iexact=cmp.lower()).exists():
        return JsonResponse({"is_taken": True})
    else:
        return JsonResponse({"is_taken": False})


@login_required(login_url="login")
def goItems(request):
    cmp = Company.objects.get(user=request.user.id)
    iData = Items.objects.filter(cid=cmp).first()
    context = {
        "cmp": cmp,
        "items": Items.objects.filter(cid=cmp),
        "item_data": iData,
        "item_transaction": Item_transactions.objects.filter(
            cid=cmp, item=iData
        ).order_by("-id"),
    }
    return render(request, "items.html", context)


@login_required(login_url="login")
def showItemData(request, id):
    cmp = Company.objects.get(user=request.user.id)
    iData = Items.objects.get(cid=cmp, id=id)
    context = {
        "cmp": cmp,
        "items": Items.objects.filter(cid=cmp),
        "item_data": iData,
        "item_transaction": Item_transactions.objects.filter(
            cid=cmp, item=iData
        ).order_by("-id"),
    }
    return render(request, "items.html", context)


@login_required(login_url="login")
def checkItemName(request):
    cmp = Company.objects.get(user=request.user.id)
    itemName = request.GET['itemName'].strip()
    if Items.objects.filter(cid = cmp, name__iexact = itemName.lower()).exists():
        return JsonResponse({'isExists':True, 'message': f"Item '{request.GET['itemName']}' already exists, try another.!"})
    return JsonResponse({'isExists':False})


@login_required(login_url="login")
def addNewItem(request):
    context = {
        "cmp": Company.objects.get(user=request.user.id),
        "itemunit": Item_units.objects.filter(
            cid=Company.objects.get(user=request.user.id)
        ),
    }
    return render(request, "additem.html", context)


@login_required(login_url="login")
def createNewItem(request):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        try:
            if request.method == "POST":
                itemName = request.POST['name'].strip()
                if Items.objects.filter(cid = cmp, name__iexact = itemName.lower()).exists():
                    messages.error(request, f"{itemName} exists, Try another.!")
                    return redirect(addNewItem)
                item = Items(
                    cid=cmp,
                    name=request.POST["name"],
                    hsn=request.POST["hsn"],
                    unit=request.POST["item_unit"],
                    tax=request.POST["tax"],
                    sale_price=request.POST["sale_price"],
                    purchase_price=request.POST["purchase_price"],
                    stock=request.POST["stock"],
                )
                item.save()

                # Opening stock transaction
                transaction = Item_transactions(
                    cid=cmp, item=item, type="Opening Stock",date = item.date, quantity=item.stock
                )
                transaction.save()

                if "next_item" in request.POST:
                    return redirect(addNewItem)
                else:
                    return redirect(goItems)
            else:
                messages.error(request, "Something went wrong, Please try again..!")
                return redirect(addNewItem)
        except Exception as e:
            print(e)
            messages.error(request, "Something went wrong, Please try again..!")
            return redirect(addNewItem)
    else:
        messages.error(request, "Something went wrong, Please try again..!")
        return redirect(addNewItem)

@login_required(login_url="login")
def deleteItem(request, id):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        item = Items.objects.get(cid=cmp, id=id)
        try:
            if Sales_items.objects.filter(item = item).exists() or Purchase_items.objects.filter(item = item).exists():
                messages.error(request, f"Item cannot be deleted because of Sales or Purchase transactions exists for `{item.name}`.")
                return redirect(showItemData, id)
            item.delete()
            messages.success(request, 'Item Deleted Successfully.!')
            return redirect(goItems)
        except Exception as e:
            print(e)
            return redirect(showItemData, id)
    return redirect("/")


@login_required(login_url="login")
def editItem(request,id):
        if request.user:
            cmp = Company.objects.get(user=request.user.id)
            try:
                itemData = Items.objects.get(cid = cmp, id = id)
                trns = Item_transactions.objects.get(cid = cmp, item = itemData, type = "Opening Stock")
                op_stock = trns.quantity
                context = {
                    'cmp':cmp,
                    'item':itemData,
                    'op_stock': op_stock,
                    "itemunit": Item_units.objects.filter(cid=Company.objects.get(user=request.user.id)),
                }
                return render(request,'edititem.html',context)
            except Exception as e:
                print(e)
                return redirect(showItemData, id)
        return redirect("/")


@login_required(login_url="login")
def editItemData(request,id):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        item = Items.objects.get(cid = cmp, id = id)
        trns = Item_transactions.objects.filter(cid = cmp , item = item.id).filter(type = 'Opening Stock').first()
        crQty = trns.quantity
        chQty = int(request.POST['stock'])
        diff = abs(crQty - chQty)
        try:
            if request.method == 'POST':
                item.name = request.POST['name']
                item.hsn = request.POST['hsn']
                item.unit = request.POST['item_unit']
                item.tax = request.POST['tax']
                item.sale_price = request.POST['sale_price']
                if chQty > crQty:
                    item.stock += diff
                elif chQty < crQty:
                    item.stock -= diff
                # item.stock = request.POST['stock']
                item.purchase_price = request.POST['purchase_price']

                item.save()

                trns.quantity = request.POST['stock']
                trns.save()

                return redirect(showItemData,id)
        except Exception as e:
            print(e)
            return redirect(editItem,id)
    return redirect('/')


@login_required(login_url="login")
def createitemunit(request):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        try:
            if request.method == "POST":
                unit = Item_units(
                    cid=cmp, symbol=request.POST["usymbol"], name=request.POST["uname"]
                )
                unit.save()
                return JsonResponse({"message": "success"})
            else:
                return JsonResponse({"message": "failed"})
        except Exception as e:
            print(e)
            return JsonResponse({"message": "failed"})
    return JsonResponse({"message": "failed"})


def getItemUnits(request):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            options = {}
            list = []
            option_objects = Item_units.objects.filter(cid=cmp)

            for item in option_objects:
                itemUnitDict = {
                    "symbol": item.symbol,
                    "name": item.name,
                }
                list.append(itemUnitDict)

            print(list)
            return JsonResponse({"units": list}, safe=False)
        except Exception as e:
            print(e)
            return JsonResponse({"message": "failed"})
    else:
        return JsonResponse({"message": "failed"})


@login_required(login_url="login")
def updateStock(request, id):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        try:
            if request.method == "POST":
                item = Items.objects.get(cid=cmp, id=id)
                if not "update_qty" in request.POST:
                    print("num===", int(request.POST["qty_update"]))
                    item.stock += int(request.POST["qty_update"])
                    item.save()

                    trns = Item_transactions(
                        cid=cmp,
                        item=item,
                        type="Add Stock",
                        date=request.POST["update_date"],
                        quantity=request.POST["qty_update"],
                    )
                    trns.save()
                    return redirect(showItemData, id)
                else:
                    print("num===", int(request.POST["qty_update"]))
                    item.stock -= int(request.POST["qty_update"])
                    item.save()

                    trns = Item_transactions(
                        cid=cmp,
                        item=item,
                        type="Reduce Stock",
                        date=request.POST["update_date"],
                        quantity=request.POST["qty_update"],
                    )
                    trns.save()
                    return redirect(showItemData, id)
        except Exception as e:
            print(e)
            return redirect(showItemData, id)
    return redirect("/")


@login_required(login_url="login")
def deleteTransaction(request, id):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        trns = Item_transactions.objects.get(cid=cmp, id=id)
        try:
            item = Items.objects.get(id = trns.item.id)
            if trns.type == "Add Stock":
                item.stock -= trns.quantity
            elif trns.type == "Reduce Stock":
                item.stock += trns.quantity
            
            item.save()
            trns.delete()
            return redirect(showItemData, trns.item.id)
        except Exception as e:
            print(e)
            return redirect(showItemData, trns.item.id)
    return redirect("/")


@login_required(login_url="login")
def editTransaction(request,id):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        try:
            trns = Item_transactions.objects.get(cid=cmp, id=id)
            context = {
                'cmp':cmp,
                'transaction':trns,
            }
            return render(request, 'edit_transaction.html',context)
        except Exception as e:
            print(e)
            return redirect(showItemData, trns.item.id)
    return redirect("/")


@login_required(login_url="login")
def editTransactionData(request, id):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        trns = Item_transactions.objects.get(cid=cmp, id=id)
        item = Items.objects.get(cid =cmp, id = trns.item.id)
        crQty = trns.quantity
        chQty = int(request.POST['quantity'])
        diff = abs(crQty - chQty)
        try:
            if request.method == 'POST':
                trns.type = request.POST['type']
                if str(request.POST['type']).lower() == 'reduce stock' and chQty > crQty:
                    item.stock -= diff
                elif str(request.POST['type']).lower() == 'reduce stock' and chQty < crQty:
                    item.stock += diff
                elif str(request.POST['type']).lower() == 'add stock' and chQty > crQty:
                    item.stock += diff
                elif str(request.POST['type']).lower() == 'add stock' and chQty < crQty:
                    item.stock -= diff

                if str(request.POST['type']).lower() == 'opening stock' and  chQty > crQty:
                    item.stock += diff
                elif str(request.POST['type']).lower() == 'opening stock' and chQty < crQty:
                    item.stock -= diff

                trns.quantity = request.POST['quantity']
                trns.date = request.POST['date']
                trns.save()
                item.save()

                return redirect(showItemData, trns.item.id)
        except Exception as e:
            print(e)
            return redirect(editTransaction, id)
    return redirect("/")



# PURCHASES
    
@login_required(login_url="login")
def goPurchases(request):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            context = {
                'cmp': cmp,
                'purchases': Purchases.objects.filter(cid = cmp)
            }
            return render(request, 'purchases.html',context)
        except Exception as e:
            print(e)
            return redirect(goDashboard)
    return redirect('/')


@login_required(login_url="login")
def addNewPurchase(request):
    if request.user:
        cmp = Company.objects.get(user = request.user.id)
        try:
            
            # Fetching last bill and assigning upcoming bill no as current + 1
            # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
            latest_bill = Purchases.objects.filter(cid = cmp).order_by('-bill_no').first()

            if latest_bill:
                last_number = int(latest_bill.bill_number)
                new_number = last_number + 1
            else:
                new_number = 1

            if DeletedPurchases.objects.filter(cid = cmp).exists():
                deleted = DeletedPurchases.objects.get(cid = cmp)
                
                if deleted:
                    while int(deleted.bill_number) >= new_number:
                        new_number+=1


            # while Purchases.objects.filter(deleted_bill_no__bill_number = new_number).exists():
            #     new_number += 1

                # model_meta = Purchases._meta
                # pk_name = model_meta.pk.name
                # table_name = model_meta.db_table
                # with connection.cursor() as cursor:
                #     cursor.execute(f"SELECT AUTO_INCREMENT FROM information_schema.TABLES WHERE TABLE_NAME = %s", [table_name])
                #     next_id = cursor.fetchone()[0]

                # cmp = Company.objects.get(user=request.user.id)


            items = Items.objects.filter( cid = cmp)
            context = {
                'cmp': cmp,
                'bill_no':new_number,
                'items': items,

            }
            return render(request, 'add_purchase.html',context)
        except Exception as e:
            print(e)
            return redirect(goDashboard)
    return redirect('/')


def getItemData(request):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            id = request.GET.get('id')

            item = Items.objects.get(name = id, cid=cmp)
            hsn = item.hsn
            pur_rate = item.purchase_price
            sale_rate = item.sale_price
            tax = item.tax
            return JsonResponse({"status":True,'id':item.id,'hsn':hsn,'pur_rate':pur_rate,'sale_rate':sale_rate, 'tax':tax})
        except Exception as e:
            print(e)
            return JsonResponse({"status":False})
    return redirect('/')


@login_required(login_url="login")
def createNewPurchase(request):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            if request.method == 'POST':
                purchase = Purchases(
                    cid = cmp,
                    bill_number = request.POST['bill_no'],
                    date = request.POST['date'],
                    party_name = request.POST['party_name'],
                    phone_number = request.POST['party_phone'],
                    gstin = request.POST['party_gstin'],
                    subtotal = request.POST['subtotal'],
                    tax = request.POST['tax'],
                    adjustment = request.POST['adjustment'],
                    total_amount = request.POST['grand_total'],
                )
                purchase.save()
                ids = request.POST.getlist("pItems[]")
                item = request.POST.getlist("item[]")
                hsn  = request.POST.getlist("hsn[]")
                qty = request.POST.getlist("qty[]")
                price = request.POST.getlist("price[]")
                tax = request.POST.getlist("tax[]")
                total = request.POST.getlist("total[]")

                pid = Purchases.objects.get( bill_no = purchase.bill_no)

                if len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(total)==len(ids) and ids and item and hsn and qty and price and tax and total:
                    mapped = zip(item,hsn,qty,price,tax,total,ids)
                    mapped = list(mapped)
                    for ele in mapped:
                        pItems = Purchase_items.objects.create(name = ele[0],hsn=ele[1],quantity=ele[2],rate=ele[3],tax=ele[4],total=ele[5],pid = pid, cid=cmp, item = Items.objects.get(cid = cmp, id = ele[6]))

                # Add purchase details in items transactions
                if len(ids) == len(qty) and ids and qty:
                    itms = zip(ids,qty)
                    trns = list(itms)
                    for itm in trns:
                        tItem = Items.objects.get(id = itm[0], cid = cmp)
                        transaction = Item_transactions.objects.create(cid = cmp, item = tItem, type = 'Purchase', date = purchase.date, quantity = itm[1], bill_number = purchase.bill_number)
                        tItem.stock += int(itm[1])
                        tItem.save()


                if 'new_purchase' in request.POST:
                    return redirect(addNewPurchase)
                return redirect(goPurchases)
        except Exception as e:
            print(e)
            messages.error(request, 'Something went wrong, Please try again.!')
            return redirect(addNewPurchase)
    return redirect('/')


@login_required(login_url="login")
def purchasesInBetween(request):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            if request.method == 'GET':
                start_date = request.GET['start_date']
                end_date = request.GET['end_date']
                purchases = Purchases.objects.filter(cid = cmp).filter(date__gte = start_date, date__lte = end_date)
                if not purchases:
                    messages.warning(request, f'No purchases found in between {start_date} to {end_date}')
                    # purchases = Purchases.objects.filter(cid = cmp)
                    return redirect(goPurchases)
                context = {
                    'cmp': cmp,
                    'purchases': purchases,
                    'start':start_date,
                    'end':end_date,
                }
                return render(request, 'purchases.html',context)
        except Exception as e:
            print(e)
            return redirect(goPurchases)
    return redirect('/')


@login_required(login_url="login")
def viewPurchaseBill(request,id):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            purchases = Purchases.objects.filter(cid = cmp)
            bill = Purchases.objects.get(cid = cmp, bill_no = id)
            items = Purchase_items.objects.filter(cid = cmp, pid = bill)
            context = {
                'cmp': cmp,
                'purchases':purchases,
                'bill': bill,
                'items':items,
            }
            return render(request, 'purchase_bill.html',context)
        except Exception as e:
            print(e)
            return redirect(goPurchases)
    return redirect('/')


@login_required(login_url="login")
def editPurchaseBill(request,id):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            bill = Purchases.objects.get(cid = cmp, bill_no = id)
            p_items = Purchase_items.objects.filter(cid = cmp, pid = bill)
            items = Items.objects.filter(cid = cmp)
            context = {
                'cmp': cmp,
                'bill': bill,
                'items':items,
                'purchase_items':p_items,
            }

            return render(request, 'edit_purchase_bill.html',context)
        except Exception as e:
            print(e)
            return redirect(viewPurchaseBill, id)
    return redirect('/')


@login_required(login_url="login")
def updatePurchaseBill(request,id):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            bill = Purchases.objects.get(cid = cmp, bill_no = id)
            bill.bill_number = request.POST['bill_no']

            bill.date = request.POST['date']
            if 'party' in request.POST:
                bill.party_name = request.POST['party_name']
                bill.phone_number = request.POST['party_phone']
                bill.gstin = request.POST['party_gstin']
            else:
                bill.party_name = ""
                bill.phone_number = ""
                bill.gstin = ""
            
            bill.subtotal = request.POST['subtotal']
            bill.tax = request.POST['tax']
            bill.adjustment = request.POST['adjustment']
            bill.total_amount = request.POST['grand_total']

            bill.save()

            ids = request.POST.getlist("pItems[]")
            item = request.POST.getlist("item[]")
            hsn  = request.POST.getlist("hsn[]")
            qty = request.POST.getlist("qty[]")
            price = request.POST.getlist("price[]")
            tax = request.POST.getlist("tax[]")
            total = request.POST.getlist("total[]")
            prchs_item_ids = request.POST.getlist("id[]")
            item_ids = [int(id) for id in prchs_item_ids]

            
            prchs_item = Purchase_items.objects.filter(pid = bill)
            object_ids = [obj.id for obj in prchs_item]

            ids_to_delete = [obj_id for obj_id in object_ids if obj_id not in item_ids]
            for id in ids_to_delete:
                purItem = Purchase_items.objects.get(cid = cmp, id = id)
                itm = Items.objects.get(cid = cmp, id = purItem.item.id)
                itm.stock -= purItem.quantity
                itm.save()
                Item_transactions.objects.filter(cid = cmp, bill_number = bill.bill_number, type = 'Purchase', item = itm).delete()

            Purchase_items.objects.filter(id__in=ids_to_delete).delete()
            
            count = Purchase_items.objects.filter(pid = bill, cid = cmp).count()
            if len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(total)==len(ids):
                try:
                    mapped=zip(item,hsn,qty,price,tax,total,item_ids,ids)
                    mapped=list(mapped)
                    
                    for ele in mapped:
                        if int(len(item))>int(count):
                            if ele[6] == 0:
                                Purchase_items.objects.create(name = ele[0], hsn=ele[1],quantity=ele[2],rate=ele[3],tax=ele[4],total=ele[5] ,pid = bill ,cid = cmp, item = Items.objects.get(cid = cmp, id = ele[7]))
                                tItem = Items.objects.get(id = ele[7], cid = cmp)
                                transaction = Item_transactions.objects.create(cid = cmp, item = tItem, type = 'Purchase', date = bill.date, quantity = ele[2], bill_number = bill.bill_number)
                                tItem.stock += int(ele[2])
                                tItem.save()
                            else:
                                Purchase_items.objects.filter( id = ele[6],cid = cmp).update(name = ele[0],hsn=ele[1],quantity=ele[2],rate=ele[3],tax=ele[4],total=ele[5], item = Items.objects.get(cid = cmp, id = ele[7]))
                                tItem = Items.objects.get(id = ele[7], cid = cmp)
                                transaction = Item_transactions.objects.get(cid =cmp, type = 'Purchase',bill_number = bill.bill_number,item = ele[7])
                                crQty = int(transaction.quantity)
                                if crQty < int(ele[2]):
                                    tItem.stock +=  abs(crQty - int(ele[2]))
                                elif crQty > int(ele[2]):
                                    tItem.stock -= abs(crQty - int(ele[2]))
                                tItem.save()
                                transaction.quantity = int(ele[2])
                                transaction.save()
                        else:
                            Purchase_items.objects.filter( id = ele[6],cid=cmp).update(name = ele[0],hsn=ele[1],quantity=ele[2],rate=ele[3],tax=ele[4],total=ele[5], item = Items.objects.get(cid = cmp, id = ele[7]))
                            tItem = Items.objects.get(id = ele[7], cid = cmp)
                            transaction = Item_transactions.objects.get(cid =cmp, type = 'Purchase',bill_number = bill.bill_number,item = ele[7])
                            crQty = int(transaction.quantity)
                            if crQty < int(ele[2]):
                                tItem.stock +=  abs(crQty - int(ele[2]))
                            elif crQty > int(ele[2]):
                                tItem.stock -= abs(crQty - int(ele[2]))
                            tItem.save()
                            transaction.quantity = int(ele[2])
                            transaction.save()
                            
                except Exception as e:
                        print(e)
                        mapped=zip(item,hsn,qty,price,tax,total,item_ids,ids)
                        mapped=list(mapped)
                        
                        for ele in mapped:
                            Purchase_items.objects.filter(id=ele[6] ,cid=cmp).update(name = ele[0],hsn=ele[1],quantity=ele[2],rate=ele[3],tax=ele[4],total=ele[5], item = Items.objects.get(cid = cmp, id = ele[7]))
                            tItem = Items.objects.get(id = ele[7], cid = cmp)
                            transaction = Item_transactions.objects.get(cid =cmp, type = 'Purchase',bill_number = bill.bill_number,item = ele[7])
                            crQty = int(transaction.quantity)
                            if crQty < int(ele[2]):
                                tItem.stock +=  abs(crQty - int(ele[2]))
                            elif crQty > int(ele[2]):
                                tItem.stock -= abs(crQty - int(ele[2]))
                            tItem.save()
                            transaction.quantity = int(ele[2])
                            transaction.save()

            return redirect(viewPurchaseBill,bill.bill_no)
        except Exception as e:
            print(e)
            return redirect(editPurchaseBill,bill.bill_no)
    return redirect('/')



@login_required(login_url="login")
def deletePurchaseBill(request, id):
    if request.user:
        try:
            cmp = Company.objects.get(user = request.user.id)
            bill = Purchases.objects.get(cid = cmp, bill_no = id)
            items = Purchase_items.objects.filter(cid = cmp, pid = bill)
            for i in items:
                purItem = Purchase_items.objects.get(cid = cmp, id = i.id)
                itm = Items.objects.get(cid = cmp, id = purItem.item.id)
                itm.stock -= purItem.quantity
                itm.save()
                Item_transactions.objects.filter(cid = cmp, bill_number = bill.bill_number, type = 'Purchase', item = itm).delete()

            Purchase_items.objects.filter(cid = cmp, pid = bill).delete()

            # Storing bill number to deleted table
            # if entry exists and lesser than the current, update and save => Only one entry per company
            if DeletedPurchases.objects.filter(cid = cmp).exists():
                deleted = DeletedPurchases.objects.get(cid = cmp)
                if deleted:
                    if int(bill.bill_number) > int(deleted.bill_number):
                        deleted.bill_number = bill.bill_number
                        deleted.save()
                
            else:
                deleted = DeletedPurchases(cid = cmp, bill_number = bill.bill_number)
                deleted.save()

            bill.delete()
            
            return redirect(goPurchases)
        except Exception as e:
            print(e)
            return redirect(viewPurchaseBill, id)
    return redirect('/')




# SALES

@login_required(login_url="login")
def goSales(request):
    if request.user:
        try:
            cmp = Company.objects.get(user = request.user.id)
            context = {
                'cmp':cmp,
                'sales':Sales.objects.filter(cid = cmp),
            }
            return render(request, 'sales.html',context)
        except Exception as e:
            print(e)
            return redirect(goDashboard)
    return redirect('/')


@login_required(login_url="login")
def addNewSale(request):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        try:

            # Fetching last bill and assigning upcoming bill no as current + 1
            # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
            latest_bill = Sales.objects.filter(cid = cmp).order_by('-bill_no').first()

            if latest_bill:
                last_number = int(latest_bill.bill_number)
                new_number = last_number + 1
            else:
                new_number = 1

            if DeletedSales.objects.filter(cid = cmp).exists():
                deleted = DeletedSales.objects.get(cid = cmp)
                
                if deleted:
                    while int(deleted.bill_number) >= new_number:
                        new_number+=1

            # model_meta = Sales._meta
            # pk_name = model_meta.pk.name
            # table_name = model_meta.db_table
            # with connection.cursor() as cursor:
            #     cursor.execute(f"SELECT AUTO_INCREMENT FROM information_schema.TABLES WHERE TABLE_NAME = %s", [table_name])
            #     next_id = cursor.fetchone()[0]

            
            items = Items.objects.filter( cid = cmp)
            context = {
                'cmp': cmp,
                'bill_no':new_number,
                'items': items,

            }
            return render(request, 'add_sale.html',context)
        except Exception as e:
            print(e)
            return redirect(goDashboard)
    return redirect('/')


@login_required(login_url="login")
def createNewSale(request):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            if request.method == 'POST':
                sale = Sales(
                    cid = cmp,
                    date = request.POST['date'],
                    bill_number = request.POST['bill_no'],
                    party_name = request.POST['party_name'],
                    phone_number = request.POST['party_phone'],
                    gstin = request.POST['party_gstin'],
                    subtotal = request.POST['subtotal'],
                    tax = request.POST['tax'],
                    adjustment = request.POST['adjustment'],
                    total_amount = request.POST['grand_total'],
                )
                sale.save()
                
                ids = request.POST.getlist('sItems[]')
                item = request.POST.getlist("item[]")
                hsn  = request.POST.getlist("hsn[]")
                qty = request.POST.getlist("qty[]")
                price = request.POST.getlist("price[]")
                tax = request.POST.getlist("tax[]")
                total = request.POST.getlist("total[]")

                sid = Sales.objects.get( bill_no = sale.bill_no)

                if len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(total)==len(ids) and ids and item and hsn and qty and price and tax and total:
                    mapped = zip(item,hsn,qty,price,tax,total,ids)
                    mapped = list(mapped)
                    for ele in mapped:
                        sItems = Sales_items.objects.create(name = ele[0],hsn=ele[1],quantity=ele[2],rate=ele[3],tax=ele[4],total=ele[5],sid = sid, cid=cmp, item = Items.objects.get(cid = cmp, id = ele[6]))
                
                # Add sales details in items transactions
                if len(ids) == len(qty) and ids and qty:
                    itms = zip(ids,qty)
                    trns = list(itms)
                    for itm in trns:
                        tItem = Items.objects.get(id = itm[0], cid = cmp)
                        transaction = Item_transactions.objects.create(cid = cmp, item = tItem, type = 'Sale', date = sale.date, quantity = itm[1], bill_number = sale.bill_number)
                        tItem.stock -= int(itm[1])
                        tItem.save()

                if 'new_sale' in request.POST:
                    return redirect(addNewSale)
                return redirect(goSales)
        except Exception as e:
            print(e)
            return redirect(addNewSale)
    return redirect('/')


@login_required(login_url="login")
def salesInBetween(request):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            if request.method == 'GET':
                start_date = request.GET['start_date']
                end_date = request.GET['end_date']
                sales = Sales.objects.filter(cid = cmp).filter(date__gte = start_date, date__lte = end_date)
                if not sales:
                    messages.warning(request, f'No sales found in between {start_date} to {end_date}')
                    # sales = Sales.objects.filter(cid = cmp)
                    return redirect(goSales)
                context = {
                    'cmp': cmp,
                    'sales': sales,
                    'start':start_date,
                    'end':end_date,
                }
                return render(request, 'sales.html',context)
        except Exception as e:
            print(e)
            return redirect(goSales)
    return redirect('/')


@login_required(login_url="login")
def viewSalesBill(request,id):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            sales = Sales.objects.filter(cid = cmp)
            bill = Sales.objects.get(cid = cmp, bill_no = id)
            items = Sales_items.objects.filter(cid = cmp, sid = bill)
            context = {
                'cmp': cmp,
                'sales':sales,
                'bill': bill,
                'items':items,
            }
            return render(request, 'sales_bill.html',context)
        except Exception as e:
            print(e)
            return redirect(goSales)
    return redirect('/')


@login_required(login_url="login")
def editSalesBill(request,id):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            bill = Sales.objects.get(cid = cmp, bill_no = id)
            s_items = Sales_items.objects.filter(cid = cmp, sid = bill)
            items = Items.objects.filter(cid = cmp)
            context = {
                'cmp': cmp,
                'bill': bill,
                'items':items,
                'sales_items':s_items,
            }

            return render(request, 'edit_sale.html',context)
        except Exception as e:
            print(e)
            return redirect(viewSalesBill, id)
    return redirect('/')



@login_required(login_url="login")
def updateSaleBill(request,id):
    if request.user:
        try:
            cmp = Company.objects.get(user=request.user.id)
            bill = Sales.objects.get(cid = cmp, bill_no = id)
            bill.bill_number = request.POST['bill_no']

            bill.date = request.POST['date']
            if 'party' in request.POST:
                bill.party_name = request.POST['party_name']
                bill.phone_number = request.POST['party_phone']
                bill.gstin = request.POST['party_gstin']
            else:
                bill.party_name = ""
                bill.phone_number = ""
                bill.gstin = ""
            
            bill.subtotal = request.POST['subtotal']
            bill.tax = request.POST['tax']
            bill.adjustment = request.POST['adjustment']
            bill.total_amount = request.POST['grand_total']

            bill.save()

            ids = request.POST.getlist('sItems[]')
            item = request.POST.getlist("item[]")
            hsn  = request.POST.getlist("hsn[]")
            qty = request.POST.getlist("qty[]")
            price = request.POST.getlist("price[]")
            tax = request.POST.getlist("tax[]")
            total = request.POST.getlist("total[]")
            sales_item_ids = request.POST.getlist("id[]")
            item_ids = [int(id) for id in sales_item_ids]

            print('item==',ids)
            print('entry==',item_ids)
            sales_item = Sales_items.objects.filter(sid = bill)
            object_ids = [obj.id for obj in sales_item]

            ids_to_delete = [obj_id for obj_id in object_ids if obj_id not in item_ids]
            for id in ids_to_delete:
                saleItem = Sales_items.objects.get(cid = cmp, id = id)
                itm = Items.objects.get(cid = cmp, id = saleItem.item.id)
                itm.stock += saleItem.quantity
                itm.save()
                Item_transactions.objects.filter(cid = cmp, bill_number = bill.bill_number, type = 'Sale', item = itm).delete()

            Sales_items.objects.filter(id__in=ids_to_delete).delete()
            
            count = Sales_items.objects.filter(sid = bill, cid = cmp).count()
            if len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(total)==len(ids):
                try:
                    mapped=zip(item,hsn,qty,price,tax,total,item_ids,ids)
                    mapped=list(mapped)
                    
                    for ele in mapped:
                        if int(len(item))>int(count):
                            if ele[6] == 0:
                                Sales_items.objects.create(name = ele[0], hsn=ele[1],quantity=ele[2],rate=ele[3],tax=ele[4],total=ele[5] ,sid = bill ,cid = cmp, item = Items.objects.get(cid = cmp, id = ele[7]))
                                tItem = Items.objects.get(id = ele[7], cid = cmp)
                                transaction = Item_transactions.objects.create(cid = cmp, item = tItem, type = 'Sale', date = bill.date, quantity = ele[2], bill_number = bill.bill_number)
                                tItem.stock -= int(ele[2])
                                tItem.save()
                            else:
                                Sales_items.objects.filter( id = ele[6],cid = cmp).update(name = ele[0],hsn=ele[1],quantity=ele[2],rate=ele[3],tax=ele[4],total=ele[5], item = Items.objects.get(cid = cmp, id = ele[7]))
                                tItem = Items.objects.get(id = ele[7], cid = cmp)
                                transaction = Item_transactions.objects.get(cid =cmp, type = 'Sale',bill_number = bill.bill_number,item = ele[7])
                                crQty = int(transaction.quantity)
                                if crQty < int(ele[2]):
                                    tItem.stock -=  abs(crQty - int(ele[2]))
                                elif crQty > int(ele[2]):
                                    tItem.stock += abs(crQty - int(ele[2]))
                                tItem.save()
                                transaction.quantity = int(ele[2])
                                transaction.save()
                        else:
                            Sales_items.objects.filter( id = ele[6],cid=cmp).update(name = ele[0],hsn=ele[1],quantity=ele[2],rate=ele[3],tax=ele[4],total=ele[5], item = Items.objects.get(cid = cmp, id = ele[7]))
                            tItem = Items.objects.get(id = ele[7], cid = cmp)
                            transaction = Item_transactions.objects.get(cid =cmp, type = 'Sale',bill_number = bill.bill_number,item = ele[7])
                            crQty = int(transaction.quantity)
                            if crQty < int(ele[2]):
                                tItem.stock -=  abs(crQty - int(ele[2]))
                            elif crQty > int(ele[2]):
                                tItem.stock += abs(crQty - int(ele[2]))
                            tItem.save()
                            transaction.quantity = int(ele[2])
                            transaction.save()
                except Exception as e:
                        print(e)
                        mapped=zip(item,hsn,qty,price,tax,total,item_ids,ids)
                        mapped=list(mapped)
                        
                        for ele in mapped:
                            Sales_items.objects.filter(id=ele[6] ,cid=cmp).update(name = ele[0],hsn=ele[1],quantity=ele[2],rate=ele[3],tax=ele[4],total=ele[5], item = Items.objects.get(cid = cmp, id = ele[7]))
                            tItem = Items.objects.get(id = ele[7], cid = cmp)
                            transaction = Item_transactions.objects.get(cid =cmp, type = 'Sale',bill_number = bill.bill_number,item = ele[7])
                            crQty = int(transaction.quantity)
                            if crQty < int(ele[2]):
                                tItem.stock -=  abs(crQty - int(ele[2]))
                            elif crQty > int(ele[2]):
                                tItem.stock += abs(crQty - int(ele[2]))
                            tItem.save()
                            transaction.quantity = int(ele[2])
                            transaction.save()


            return redirect(viewSalesBill,bill.bill_no)
        except Exception as e:
            print(e)
            return redirect(editSalesBill,bill.bill_no)
        
    return redirect('/')



@login_required(login_url="login")
def deleteSaleBill(request, id):
    if request.user:
        try:
            cmp = Company.objects.get(user = request.user.id)
            bill = Sales.objects.get(cid = cmp, bill_no = id)
            items = Sales_items.objects.filter(cid = cmp, sid = bill)
            for i in items:
                saleItem = Sales_items.objects.get(cid = cmp, id = i.id)
                itm = Items.objects.get(cid = cmp, id = saleItem.item.id)
                itm.stock += saleItem.quantity
                itm.save()
                Item_transactions.objects.filter(cid = cmp, bill_number = bill.bill_number, type = 'Sale', item = itm).delete()
            Sales_items.objects.filter(cid = cmp, sid = bill).delete()
            
            # Storing bill number to deleted table
            # if entry exists and lesser than the current, update and save => Only one entry per company

            if DeletedSales.objects.filter(cid = cmp).exists():
                deleted = DeletedSales.objects.get(cid = cmp)
                if deleted:
                    if int(bill.bill_number) > int(deleted.bill_number):
                        deleted.bill_number = bill.bill_number
                        deleted.save()
                
            else:
                deleted = DeletedSales(cid = cmp, bill_number = bill.bill_number)
                deleted.save()

            bill.delete()
            # items.delete()
            return redirect(goSales)
        except Exception as e:
            print(e)
            return redirect(viewSalesBill, id)
    return redirect('/')



@login_required(login_url="login")
def salesBillPdf(request,id):
    if request.user:
        cmp = Company.objects.get( user = request.user.id)
        bill = Sales.objects.get(cid = cmp, bill_no = id)
        items = Sales_items.objects.filter(cid = cmp, sid = bill)
    
        total = bill.total_amount
        words_total = num2words(total)
    
    context = {'bill': bill, 'cmp': cmp,'items':items, 'total':words_total}
    
    template_path = 'sales_bill_pdf.html'
    fname = 'Bill_'+str(bill.bill_no)
    # return render(request, 'sales_bill_pdf.html',context)
    # Create a Django response object, and specify content_type as pdftemp_creditnote
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] =f'attachment; filename = sales_bill_{fname}.pdf'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funny view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response



@login_required(login_url='login')
def goStockReports(request):
    if request.user:
         cmp = Company.objects.get(user = request.user.id)
         try:
            stockList = []
            items = Items.objects.filter(cid = cmp)
            
            for item in items:
                stockIn = 0
                stockOut = 0
                for i in Item_transactions.objects.filter(cid = cmp, item = item.id).filter(type = 'Purchase'):
                    stockIn += i.quantity

                for i in Item_transactions.objects.filter(cid = cmp, item = item.id).filter(type = 'Sale'):
                    stockOut += i.quantity

                dict = {
                    'name':item.name,
                    'stockIn':stockIn,
                    'stockOut':stockOut,
                    'balance':item.stock
                }
                
                stockList.append(dict)

            context = {
                'cmp':cmp,
                'items':items,
                'stock':stockList,
            }
            return render(request, 'stock_report.html',context)
         except Exception as e:
             print(e)
             return redirect(goDashboard)
    return redirect('/')


def itemStockReport(request,id):
    if request.user:
        cmp = Company.objects.get(user = request.user.id)
        try:
            stockList = []
            item = Items.objects.get(cid = cmp, id = id)
            
            stockIn = 0
            stockOut = 0
            for i in Item_transactions.objects.filter(cid = cmp, item = item).filter(type = 'Purchase'):
                stockIn += i.quantity

            for i in Item_transactions.objects.filter(cid = cmp, item = item).filter(type = 'Sale'):
                stockOut += i.quantity

            dict = {
                'name':item.name,
                'stockIn':stockIn,
                'stockOut':stockOut,
                'balance':item.stock
            }
            
            stockList.append(dict)

            context = {
                'cmp':cmp,
                'items':Items.objects.filter(cid = cmp),
                'stock':stockList,
                'balance':item.stock,
            }
            return render(request, 'stock_report.html',context)
        except Exception as e:
             print(e)
             return redirect(goStockReports)
    return redirect('/')


def shareSalesBillToEmail(request,id):
    if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                # print(emails_list)

                cmp = Company.objects.get( user = request.user.id)
                bill = Sales.objects.get(cid = cmp, bill_no = id)
                items = Sales_items.objects.filter(cid = cmp, sid = bill)
            
                total = bill.total_amount
                words_total = num2words(total)
            
                context = {'bill': bill, 'cmp': cmp,'items':items, 'total':words_total}
                template_path = 'sales_bill_pdf.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                filename = f'Sales Bill - {bill.bill_number}.pdf'
                subject = f"SALES BILL - {bill.bill_number}"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached SALES BILL - Bill-{bill.bill_number}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.phone_number}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                messages.success(request, 'Bill has been shared via email successfully..!')
                return redirect(viewSalesBill,id)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(viewSalesBill, id)
        

def shareStockReportsToEmail(request):
    if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                # print(emails_list)

                cmp = Company.objects.get( user = request.user.id)

                excelfile = BytesIO()
                workbook = Workbook()
                workbook.remove(workbook.active)
                worksheet = workbook.create_sheet(title='Stock Reports', index=1)

                stockList = []
                items = Items.objects.filter(cid = cmp)
                
                for item in items:
                    stockIn = 0
                    stockOut = 0
                    for i in Item_transactions.objects.filter(cid = cmp, item = item.id).filter(type = 'Purchase'):
                        stockIn += i.quantity

                    for i in Item_transactions.objects.filter(cid = cmp, item = item.id).filter(type = 'Sale'):
                        stockOut += i.quantity

                    dict = {
                        'name':item.name,
                        'stockIn':stockIn,
                        'stockOut':stockOut,
                        'balance':item.stock
                    }
                    stockList.append(dict)

                columns = ['#', 'Item', 'Stock In', 'Stock Out', 'Balance']
                row_num = 1

                # Assign the titles for each cell of the header
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    cell.font = Font(bold=True)
                
                # Iterate through all coins
                print('stock===', stockList)
                sl_no = 0
                for _, bill in enumerate(stockList, 1):
                    print('bill====',bill)
                    row_num += 1
                    sl_no+=1
                    # Define the data for each cell in the row
                    name,stockin,stockout,bal = (bill.get(key) for key in ['name', 'stockIn', 'stockOut', 'balance'])
                    row = [
                        sl_no,
                        name,
                        stockin,
                        stockout,
                        bal,
                    ]

                    print('ROW=========')
                    print(row)

                    # Assign the data for each cell of the row
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.protection = Protection(locked=True)
                workbook.save(excelfile)
                mail_subject = f'Stock Reports - {date.today()}'
                message = f"Hi,\nPlease find the STOCK REPORTS file attached. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.phone_number}"
                message = EmailMessage(mail_subject, message, settings.EMAIL_HOST_USER, emails_list)
                message.attach(f'Stock Reports-{date.today()}.xlsx', excelfile.getvalue(), 'application/vnd.ms-excel')
                message.send(fail_silently=False)

                messages.success(request, 'Stock Report has been shared via email successfully..!')
                return redirect(goStockReports)
        except Exception as e:
            print(e)
            return redirect(goStockReports)


def changeTrialStatus(request, status):
    if request.user:
        trial = ClientTrials.objects.get(user = request.user)
        trial.subscribe_status = status
        trial.save()
        # return HttpResponse('<script>alert("Success.!");window.history.back();</script>')
        messages.success(request,'Success.!')
        return redirect(goDashboard)
    
    
    
@login_required(login_url='login')
def sales_report(request):
    if request.user:
        cmp = Company.objects.get(user = request.user.id)
        context = {
            'cmp':cmp,
            'sales':Sales.objects.filter(cid = cmp),
        }
        return render(request, 'sales_report.html',context)
    
    
def shareSalesReportsToEmail(request):
    if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                # print(emails_list)

                cmp = Company.objects.get( user = request.user.id)
                invoices = Sales.objects.filter(cid=cmp)


                excelfile = BytesIO()
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = 'Sales Reports'

                # Write headers
                headers = ['#', 'Date', 'Invoice Number', 'Party Name', 'Amount']
                for col_num, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col_num, value=header)

                # Write sales invoices data
                for idx, invoice in enumerate(invoices, start=2):
                    worksheet.cell(row=idx, column=1, value=idx - 1)  # Index
                    worksheet.cell(row=idx, column=2, value=invoice.date)  # Date
                    worksheet.cell(row=idx, column=3, value=invoice.bill_number)  # Invoice Number
                    worksheet.cell(row=idx, column=4, value=invoice.party_name)  # Party Name
                    worksheet.cell(row=idx, column=5, value=invoice.total_amount)  # Amount

                

                # Save workbook to BytesIO object
                workbook.save(excelfile)
                mail_subject = f'Sales Reports - {date.today()}'
                message = f"Hi,\nPlease find the ALES REPORTS file attached. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.phone_number}"
                message = EmailMessage(mail_subject, message, settings.EMAIL_HOST_USER, emails_list)
                message.attach(f'Sales Reports-{date.today()}.xlsx', excelfile.getvalue(), 'application/vnd.ms-excel')
                message.send(fail_silently=False)

                messages.success(request, 'Sales Report has been shared via email successfully..!')
                return redirect(sales_report)
        except Exception as e:
            print(e)
            messages.error(request, 'An error occurred while sharing the sales report via email.')

            return redirect(sales_report)
        
def salesreport_graph(request):
    if request.user:
        cmp = Company.objects.get(user = request.user.id)
        current_year = datetime.now().year


        monthly_sales_data = defaultdict(int)
        for month in range(1, 13):
            monthly_sales_data[month] = (
                Sales.objects
                .filter(date__month=month, date__year=current_year,cid=cmp)
                .aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
            )

        # Retrieve yearly sales data
        current_year = datetime.now().year
        yearly_sales_data = defaultdict(int)
        for year in range(2022, current_year + 1):
            yearly_sales_data[year] = (
                Sales.objects
                .filter(date__year=year,cid=cmp)
                .aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
            )

        # Prepare data for chart
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        monthly_labels = [f"{month_names[month - 1]} {current_year}" for month in range(1, 13)]
        monthly_sales = [monthly_sales_data[month] for month in range(1, 13)]

        yearly_labels = [str(year) for year in range(2014, current_year + 1)]
        yearly_sales = [yearly_sales_data[year] for year in range(2014, current_year + 1)]

        # Prepare data for chart
        chart_data = {'monthly_labels': monthly_labels, 'monthly_sales': monthly_sales,
                    'yearly_labels': yearly_labels, 'yearly_sales': yearly_sales}
        return render(request, 'saleschart.html', {'chart_data': chart_data,'cmp':cmp,})